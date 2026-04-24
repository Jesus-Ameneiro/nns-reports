"""
NNS Report Writer - fills MCC and CS templates with processed data.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date


# ---------------------------------------------------------------------------
# MCC FOOTER CONSTANTS
# Code owns the MCC Data sheet footer entirely.
# The template only needs to supply rows 1–31 (header + 18 data rows).
# Update these values when Trimble provides new contact information.
# ---------------------------------------------------------------------------

MCC_NOTE_GAP    = 5   # rows between last data row and the Nota row
MCC_CONTACT_GAP = 8   # rows between last data row and first contact line

MCC_FOOTER_NOTE = (
    "Nota: El presente documento contiene información confidencial "
    "y se proporciona exclusivamente dentro del marco de License Compliance."
)

# (text, bold) — one tuple per contact block line
MCC_CONTACT_LINES = [
    ("XXXXXXXXXX",                                    True),   # specialist name
    ("Especialista en Resolución",                    False),
    ("(xx) xxxx - xxxx",                              False),  # phone
    ("xxxxx@ruvixx.com",                              False),  # email
    ("",                                              False),
    ("425 Page Mill Rd, Suite 200, Palo Alto, 94306", False),  # address
]


# ---------------------------------------------------------------------------
# UTILITIES
# ---------------------------------------------------------------------------

def detect_template_type(wb):
    return 'MCC' if 'LC Summary' in wb.sheetnames else 'CS'

def detect_summary_sheet(wb):
    for name in ['LC Summary', 'Summary', 'New Template']:
        if name in wb.sheetnames:
            return name
    return wb.sheetnames[0]

def find_col_by_header(ws, header_row, header_name):
    for cell in ws[header_row]:
        if cell.value and str(cell.value).strip().lower() == header_name.strip().lower():
            return cell.column
    return None

def safe_set(ws, row, col, value):
    """Set cell value, skipping merged-cell slaves silently."""
    try:
        ws.cell(row=row, column=col).value = value
    except AttributeError:
        pass

def col_all_dash(ws, col_idx, data_start_row, data_end_row):
    for r in range(data_start_row, data_end_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v is not None and str(v).strip() not in ('', '-'):
            return False
    return True

def format_date(d):
    if d is None:
        return '-'
    if hasattr(d, 'strftime'):
        return d.strftime('%Y-%m-%d')
    return str(d)


# ---------------------------------------------------------------------------
# MCC TEMPLATE FILLER
# ---------------------------------------------------------------------------

def _write_mcc_footer(ws_data, last_data_row):
    """
    Write the MCC Data sheet footer at a position computed from last_data_row.
    Called for every run — regardless of machine count — so the footer is
    always correctly placed and never overwritten by machine data.

    Writes:
      • Nota row (merged A:H) at last_data_row + MCC_NOTE_GAP
      • Contact block lines starting at last_data_row + MCC_CONTACT_GAP
    """
    from openpyxl.styles import Font, Alignment

    note_row    = last_data_row + MCC_NOTE_GAP
    contact_row = last_data_row + MCC_CONTACT_GAP

    # ── Note cell ─────────────────────────────────────────────────────────
    cell            = ws_data.cell(note_row, 1)
    cell.value      = MCC_FOOTER_NOTE
    cell.font       = Font(name='Calibri', size=9, italic=True, color='404040')
    cell.alignment  = Alignment(wrap_text=True, vertical='top')

    ws_data.merge_cells(
        start_row=note_row, start_column=1,
        end_row=note_row,   end_column=8,
    )

    # ── Contact block ──────────────────────────────────────────────────────
    for offset, (text, bold) in enumerate(MCC_CONTACT_LINES):
        cell           = ws_data.cell(contact_row + offset, 1)
        cell.value     = text if text else None
        cell.font      = Font(name='Calibri', size=9, bold=bold)
        cell.alignment = Alignment(wrap_text=False)


def fill_mcc(wb, rows, globals_data, case_ids, entity_name, country):
    ws_summary = wb['LC Summary']
    ws_data    = wb['Data']

    # ---- LC Summary (footer stays at fixed template positions: 20, 23-26, 28) ----
    ws_summary['B8']  = ', '.join(case_ids)
    ws_summary['B9']  = entity_name
    ws_summary['A14'] = country
    ws_summary['B14'] = globals_data['total_machines']
    ws_summary['C14'] = globals_data['total_users']
    ws_summary['D14'] = globals_data['versions_str']
    ws_summary['E14'] = globals_data['total_events']
    ws_summary['F14'] = globals_data['total_licenses']
    ws_summary['G14'] = globals_data['years_of_use']
    ws_summary['H14'] = globals_data['period']
    ws_summary['B16'] = globals_data['total_machines']
    ws_summary['C16'] = globals_data['total_users']
    ws_summary['D16'] = globals_data['versions_str']
    ws_summary['E16'] = globals_data['total_events']
    ws_summary['F16'] = globals_data['total_licenses']
    ws_summary['G16'] = globals_data['years_of_use']

    # Check for COMPUTER DOMAIN column in summary header row
    SUMMARY_HEADER_ROW = 13
    comp_domain_col_summary = find_col_by_header(ws_summary, SUMMARY_HEADER_ROW, 'COMPUTER DOMAIN')

    # ---- Data sheet ----
    DATA_HEADER_ROW  = 13
    DATA_START_ROW   = 14
    TEMPLATE_DATA_ROWS = 18   # template has 18 pre-bordered rows (14-31)

    col_map = {}
    for cell in ws_data[DATA_HEADER_ROW]:
        if cell.value:
            col_map[str(cell.value).strip()] = cell.column

    mcc_col_order = [
        ('Active MAC',             'active_mac'),
        ('# Licenses',             'license_count'),
        ('Products',               'product'),
        ('First Event',            'first_event'),
        ('Last Event',             'last_event'),
        ('Event Types',            'event_type'),
        ('Computer Domains',       'computer_domain'),
        ('Version',                'version'),
        ('IP Country',             'ip_country'),
        ('Hostname',               'hostname'),
        ('Username',               'username'),
        ('Client Email Addresses', 'client_email'),
    ]

    n_rows = len(rows)
    n_template_cols = max(col_map.values()) if col_map else 8

    # Template layout constants
    TEMPLATE_LAST_DATA_ROW = DATA_START_ROW + TEMPLATE_DATA_ROWS - 1  # row 31

    last_data_row = DATA_START_ROW + n_rows - 1
    excess        = TEMPLATE_DATA_ROWS - n_rows

    # ── Clear old template footer zone ────────────────────────────────────
    # Must happen BEFORE writing data rows so that data written into rows
    # 32+ is not subsequently wiped. Merges cleared first to avoid the
    # AttributeError raised when assigning to a merged-cell slave.
    ws_data.merged_cells.ranges.clear()

    for r in range(TEMPLATE_LAST_DATA_ROW + 1, TEMPLATE_LAST_DATA_ROW + 25):
        for c in range(1, n_template_cols + 1):
            try:
                ws_data.cell(r, c).value = None
            except AttributeError:
                pass  # residual merged slave — safe to skip after range.clear()

    # ── Write data rows ────────────────────────────────────────────────────
    # Copy template styles to any row beyond the 18 pre-bordered template rows.
    for idx, row in enumerate(rows):
        r = DATA_START_ROW + idx
        if idx >= TEMPLATE_DATA_ROWS:
            _copy_row_style(ws_data, DATA_START_ROW, r, n_template_cols)
        for header, field in mcc_col_order:
            col_idx = col_map.get(header)
            if col_idx is None:
                continue
            val = row.get(field, '-')
            if val is None:
                val = '-'
            # Write dates as datetime objects to preserve Excel date formatting
            if field in ('first_event', 'last_event') and val != '-':
                try:
                    from datetime import datetime, date as date_type
                    if isinstance(val, date_type) and not isinstance(val, datetime):
                        val = datetime(val.year, val.month, val.day)
                except Exception:
                    val = format_date(val)
            safe_set(ws_data, r, col_idx, val)

    # ── Row count adjustment ───────────────────────────────────────────────
    if excess > 0:
        # Fewer machines than the 18 template rows: delete the empty surplus.
        # After deletion the footer is written at the correct position below.
        ws_data.delete_rows(DATA_START_ROW + n_rows, excess)

    # ── Write footer at the computed position ─────────────────────────────
    # Always done in code — works for 1 machine or 1 000+ machines.
    _write_mcc_footer(ws_data, last_data_row)

    data_end_row = last_data_row

    # Fix stale cell-level hyperlink refs after delete_rows
    for row in ws_data.iter_rows():
        for cell in row:
            if cell.hyperlink and cell.hyperlink.ref:
                correct_ref = f'{get_column_letter(cell.column)}{cell.row}'
                if cell.hyperlink.ref != correct_ref:
                    cell.hyperlink.ref = correct_ref

    # ---- Column deletion ----
    # Re-read col_map after potential column deletions from Computer Domains check
    cols_to_delete = []
    for header in ['Computer Domains', 'Client Email Addresses']:
        col_idx = col_map.get(header)
        if col_idx and col_all_dash(ws_data, col_idx, DATA_START_ROW, data_end_row):
            cols_to_delete.append(col_idx)

    if col_map.get('Computer Domains') in cols_to_delete and comp_domain_col_summary:
        ws_summary.delete_cols(comp_domain_col_summary)

    for col_idx in sorted(cols_to_delete, reverse=True):
        ws_data.delete_cols(col_idx)

    return wb


# ---------------------------------------------------------------------------
# CS TEMPLATE FILLER
# ---------------------------------------------------------------------------

def _fix_merged_cells_after_row_deletion(ws, first_deleted_row, count=1):
    """
    openpyxl's delete_rows() shifts cell contents but does NOT update merged
    cell range coordinates. Rebuild the registry, decrementing any row number
    >= first_deleted_row by count (number of rows deleted).
    """
    old_ranges = [
        (mc.min_row, mc.min_col, mc.max_row, mc.max_col)
        for mc in ws.merged_cells.ranges
    ]
    ws.merged_cells.ranges.clear()
    for (min_row, min_col, max_row, max_col) in old_ranges:
        # Drop merges entirely within the deleted range
        if min_row >= first_deleted_row and max_row < first_deleted_row + count:
            continue
        # Keep merges entirely above the deleted range unchanged
        if max_row < first_deleted_row:
            ws.merge_cells(start_row=min_row, start_column=min_col,
                           end_row=max_row, end_column=max_col)
            continue
        # Shift rows at or below the first deleted row
        if min_row >= first_deleted_row:
            min_row -= count
        if max_row >= first_deleted_row:
            max_row -= count
        if min_row < 1 or max_row < 1 or min_row > max_row:
            continue
        ws.merge_cells(start_row=min_row, start_column=min_col,
                       end_row=max_row, end_column=max_col)


def _fix_image_anchors_after_row_deletion(ws, deleted_row_0based):
    """Shift image anchors >= deleted_row_0based down by 1 (0-based rows)."""
    for img in ws._images:
        try:
            anchor = img.anchor
            if hasattr(anchor, '_from'):
                if anchor._from.row >= deleted_row_0based:
                    anchor._from.row -= 1
            if hasattr(anchor, 'to') and anchor.to is not None:
                if anchor.to.row >= deleted_row_0based:
                    anchor.to.row -= 1
        except Exception:
            pass


def _fix_image_anchors_after_rows_deletion(ws, first_deleted_0based, count):
    """Shift image anchors >= first_deleted_0based down by count (0-based rows)."""
    for img in ws._images:
        try:
            anchor = img.anchor
            if hasattr(anchor, '_from'):
                if anchor._from.row >= first_deleted_0based:
                    anchor._from.row -= count
            if hasattr(anchor, 'to') and anchor.to is not None:
                if anchor.to.row >= first_deleted_0based:
                    anchor.to.row -= count
        except Exception:
            pass


def _fix_merged_cells_after_col_deletion(ws, deleted_col_1based):
    """
    openpyxl's delete_cols() does not update merged cell col coordinates.
    Decrement any col index >= deleted_col_1based by 1.
    Merges entirely within one cell or that span the deleted col are dropped.
    """
    old_ranges = [
        (mc.min_row, mc.min_col, mc.max_row, mc.max_col)
        for mc in ws.merged_cells.ranges
    ]
    ws.merged_cells.ranges.clear()
    for (min_row, min_col, max_row, max_col) in old_ranges:
        # Drop merges entirely on the deleted column
        if min_col == deleted_col_1based and max_col == deleted_col_1based:
            continue
        # Shift cols >= deleted col left by 1
        if min_col >= deleted_col_1based:
            min_col -= 1
        if max_col >= deleted_col_1based:
            max_col -= 1
        if min_col > max_col:
            continue
        ws.merge_cells(start_row=min_row, start_column=min_col,
                       end_row=max_row, end_column=max_col)


def _fix_image_anchors_after_col_deletion(ws, deleted_col_0based):
    """Shift image anchor cols >= deleted_col_0based left by 1 (0-based cols)."""
    for img in ws._images:
        try:
            anchor = img.anchor
            if hasattr(anchor, '_from'):
                if anchor._from.col >= deleted_col_0based:
                    anchor._from.col -= 1
            if hasattr(anchor, 'to') and anchor.to is not None:
                if anchor.to.col >= deleted_col_0based:
                    anchor.to.col -= 1
        except Exception:
            pass


def _copy_row_style(ws, src_row, dst_row, max_col):
    """Copy cell styles (font, border, fill, alignment, number_format) from src_row to dst_row."""
    from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, numbers
    import copy
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        try:
            dst.font      = copy.copy(src.font)
            dst.border    = copy.copy(src.border)
            dst.fill      = copy.copy(src.fill)
            dst.alignment = copy.copy(src.alignment)
            if src.number_format and src.number_format != 'General':
                dst.number_format = src.number_format
        except Exception:
            pass


def fill_cs(wb, rows, globals_data, case_ids, entity_name, country):
    summary_name = detect_summary_sheet(wb)
    ws_summary   = wb[summary_name]
    ws_data      = wb['Data']

    # ---- Determine if Computer Domain row should be deleted ----
    # Template row 12 = 'Computer Domain'.
    # When no domain data is present, delete that row entirely so all
    # subsequent rows shift up by 1, matching the expected output exactly.
    all_comp_domains = set()
    for row in rows:
        cd = row.get('computer_domain', '-')
        if cd and cd != '-':
            for d in cd.split(','):
                d = d.strip()
                if d:
                    all_comp_domains.add(d)

    delete_domain_row = len(all_comp_domains) == 0
    COMP_DOMAIN_ROW   = 12

    if delete_domain_row:
        # 1. Delete the row
        ws_summary.delete_rows(COMP_DOMAIN_ROW)
        # 2. Fix merged cell coordinates (openpyxl bug: they don't shift automatically)
        _fix_merged_cells_after_row_deletion(ws_summary, COMP_DOMAIN_ROW)
        # 3. Fix stale cell-level hyperlink refs in summary sheet after row deletion
        for summ_row in ws_summary.iter_rows():
            for cell in summ_row:
                if cell.hyperlink and cell.hyperlink.ref:
                    correct_ref = f'{get_column_letter(cell.column)}{cell.row}'
                    if cell.hyperlink.ref != correct_ref:
                        cell.hyperlink.ref = correct_ref
        # 4. Fix image anchor rows in summary sheet (deleted row = 12, 0-based = 11)
        _fix_image_anchors_after_row_deletion(ws_summary, COMP_DOMAIN_ROW - 1)
        # Row mappings after deletion (each was original_row - 1)
        ver_row  = 12   # Version row  (was 13)
        yofu_row = 13   # Years of Use (was 14)
        vals_row = 18   # Machines/Users/Events values (was 19)
        lic_row  = 30   # Licensed copies value row (was 31)
        # 3. Update price formula: it referenced C31 which is now C30
        for r in range(25, 35):
            cell = ws_summary.cell(r, 1)
            if cell.value and 'C31' in str(cell.value):
                cell.value = str(cell.value).replace('C31', 'C30')
                break
    else:
        ws_summary['B12'] = ', '.join(sorted(all_comp_domains))
        ver_row  = 13
        yofu_row = 14
        vals_row = 19
        lic_row  = 31

    # ---- Summary sheet: fill all data fields ----
    ws_summary['B9']  = ', '.join(case_ids)
    ws_summary['B10'] = country
    ws_summary['B11'] = entity_name

    # Version string → B{ver_row} (master of B:E merge)
    # Total Versions  → G{ver_row} (free cell, not merged)
    # Single-year versions are written as integers to match Excel native format
    versions_val = globals_data['versions_str']
    try:
        if ',' not in str(versions_val):
            versions_val = int(versions_val)
    except (ValueError, TypeError):
        pass
    ws_summary.cell(ver_row, 2).value = versions_val
    ws_summary.cell(ver_row, 7).value = globals_data['total_versions']

    # Years of Use → B{yofu_row}
    # Period       → D{yofu_row} (master of D:G merge)
    ws_summary.cell(yofu_row, 2).value = globals_data['years_of_use']
    ws_summary.cell(yofu_row, 4).value = globals_data['period']

    # Machines / Users / Events numeric values
    # B, D, G are masters of their merged spans in the values row
    safe_set(ws_summary, vals_row, 2, globals_data['total_machines'])
    safe_set(ws_summary, vals_row, 4, globals_data['total_users'])
    safe_set(ws_summary, vals_row, 7, globals_data['total_events'])

    # Licensed copies (feeds the price formula)
    safe_set(ws_summary, lic_row, 3, globals_data['total_licenses'])

    # ---- Data sheet: header rows ----
    safe_set(ws_data, 6, 2,  ', '.join(case_ids))
    safe_set(ws_data, 7, 2,  entity_name)
    safe_set(ws_data, 8, 2,  country)
    # DATE: label at col 10, current date at col 11 with mm-dd-yy format
    from datetime import datetime as _dt
    safe_set(ws_data, 7, 10, 'DATE:')
    today_cell = ws_data.cell(7, 11)
    today_cell.value = _dt.today().replace(hour=0, minute=0, second=0, microsecond=0)
    today_cell.number_format = 'mm-dd-yy'

    DATA_HEADER_ROW    = 11
    DATA_START_ROW     = 12
    TEMPLATE_DATA_ROWS = 12   # 12 pre-bordered rows (12-23) in the blank template

    col_map = {}
    for cell in ws_data[DATA_HEADER_ROW]:
        if cell.value:
            col_map[str(cell.value).strip()] = cell.column

    cs_col_order = [
        ('Products',               'product'),
        ('Version',                'version'),
        ('Event Types',            'event_type'),
        ('Active MAC',             'active_mac'),
        ('# Licenses',             'license_count'),
        ('First Event',            'first_event'),
        ('Last Event',             'last_event'),
        ('Computer Domains',       'computer_domain'),
        ('IP Country',             'ip_country'),
        ('Hostname',               'hostname'),
        ('Username',               'username'),
        ('Client Email Addresses', 'client_email'),
    ]

    n_rows = len(rows)

    # Write machine data rows — dates as datetime objects (preserves template formatting)
    STYLE_SRC_ROW = DATA_START_ROW  # template data row to copy styles from
    n_template_cols = max(col_map.values()) if col_map else 12
    for idx, row in enumerate(rows):
        r = DATA_START_ROW + idx
        # For rows beyond the template pre-bordered range, copy styles from template row
        if idx >= TEMPLATE_DATA_ROWS:
            _copy_row_style(ws_data, STYLE_SRC_ROW, r, n_template_cols)
        for header, field in cs_col_order:
            col_idx = col_map.get(header)
            if col_idx is None:
                continue
            val = row.get(field, '-')
            if val is None:
                val = '-'
            if field in ('first_event', 'last_event') and val != '-':
                try:
                    from datetime import datetime, date as date_type
                    if isinstance(val, date_type) and not isinstance(val, datetime):
                        val = datetime(val.year, val.month, val.day)
                except Exception:
                    val = format_date(val)
            safe_set(ws_data, r, col_idx, val)

    # Apply correct date format to date columns before writing is complete
    date_fmt = 'yyyy\\-mm\\-dd'
    date_fields = ['First Event', 'Last Event']
    for df in date_fields:
        col_idx = col_map.get(df)
        if col_idx:
            for r in range(DATA_START_ROW, DATA_START_ROW + n_rows):
                ws_data.cell(r, col_idx).number_format = date_fmt

    # Template has TEMPLATE_DATA_ROWS pre-bordered rows.
    # If n_rows < template: delete excess rows (footer shifts up).
    # If n_rows > template: extra rows beyond template must push footer down.
    excess = TEMPLATE_DATA_ROWS - n_rows
    TEMPLATE_LAST_DATA_ROW = DATA_START_ROW + TEMPLATE_DATA_ROWS - 1  # row 23
    # Gap from last template data row to footer note and image in template
    NOTE_ROW_IN_TEMPLATE  = 32   # template row for the footer note in Data sheet
    IMAGE_ROW_IN_TEMPLATE = 31   # template row (1-based) for footer image
    NOTE_GAP  = NOTE_ROW_IN_TEMPLATE  - TEMPLATE_LAST_DATA_ROW  # 9
    IMAGE_GAP = IMAGE_ROW_IN_TEMPLATE - TEMPLATE_LAST_DATA_ROW  # 8

    if excess > 0:
        # Fewer machines than template rows: delete excess, footer shifts up naturally
        _fix_image_anchors_after_rows_deletion(ws_data, DATA_START_ROW + n_rows - 1, excess)
        ws_data.delete_rows(DATA_START_ROW + n_rows, excess)
        _fix_merged_cells_after_row_deletion(ws_data, DATA_START_ROW + n_rows, excess)
    elif excess < 0:
        # More machines than template rows: footer stays at original template position
        # but machine data overwrites it. Explicitly push footer below last data row.
        extra_rows = -excess  # how many rows beyond template
        data_end = DATA_START_ROW + n_rows - 1

        # Write footer note at correct position
        note_row  = data_end + NOTE_GAP
        image_row = data_end + IMAGE_GAP  # 0-based = image_row - 1

        # Re-write note text (template note at row 32 was overwritten by machine data)
        note_text = ws_data.cell(NOTE_ROW_IN_TEMPLATE, 1).value
        if not note_text or 'Note:' not in str(note_text):
            note_text = ('Note: This document contains confidential information and is '
                         'provided exclusively within the framework of License Compliance.')
        ws_data.cell(note_row, 1).value = note_text

        # Shift footer merges (A32:J32 and D28:M28) to correct rows
        old_ranges = [
            (mc.min_row, mc.min_col, mc.max_row, mc.max_col)
            for mc in ws_data.merged_cells.ranges
        ]
        ws_data.merged_cells.ranges.clear()
        for (min_row, min_col, max_row, max_col) in old_ranges:
            if min_row >= NOTE_ROW_IN_TEMPLATE - NOTE_GAP:  # footer area rows (>=28)
                min_row += extra_rows
                max_row += extra_rows
            ws_data.merge_cells(start_row=min_row, start_column=min_col,
                                 end_row=max_row, end_column=max_col)

        # Shift footer image anchor to correct row (0-based)
        target_img_row_0based = image_row - 1  # 0-based
        for img in ws_data._images:
            try:
                if img.anchor._from.row >= IMAGE_ROW_IN_TEMPLATE - 1:  # footer image area
                    img.anchor._from.row = target_img_row_0based
            except Exception:
                pass

    data_end_row = DATA_START_ROW + n_rows - 1

    # Fix stale cell-level hyperlink refs after delete_rows
    for data_row in ws_data.iter_rows():
        for cell in data_row:
            if cell.hyperlink and cell.hyperlink.ref:
                correct_ref = f'{get_column_letter(cell.column)}{cell.row}'
                if cell.hyperlink.ref != correct_ref:
                    cell.hyperlink.ref = correct_ref

    # ---- Column deletion ----
    cols_to_delete = []
    for header in ['Computer Domains', 'Client Email Addresses']:
        col_idx = col_map.get(header)
        if col_idx and col_all_dash(ws_data, col_idx, DATA_START_ROW, data_end_row):
            cols_to_delete.append(col_idx)

    for col_idx in sorted(cols_to_delete, reverse=True):
        ws_data.delete_cols(col_idx)
        # Fix image anchor columns (openpyxl doesn't update them after delete_cols)
        _fix_image_anchors_after_col_deletion(ws_data, col_idx - 1)  # convert to 0-based
        # Fix merged cell col ranges (openpyxl doesn't update them after delete_cols)
        _fix_merged_cells_after_col_deletion(ws_data, col_idx)

    return wb


# ---------------------------------------------------------------------------
# MAIN ENTRY POINT
# ---------------------------------------------------------------------------

def fill_template(template_wb, rows, globals_data, case_ids, entity_name, country):
    wb            = template_wb
    template_type = detect_template_type(wb)
    if template_type == 'MCC':
        wb = fill_mcc(wb, rows, globals_data, case_ids, entity_name, country)
    else:
        wb = fill_cs(wb, rows, globals_data, case_ids, entity_name, country)
    return wb, template_type


def patch_and_save(wb, output_buffer):
    """Save workbook, patching style alignment corruption from delete_cols."""
    max_align = len(wb._alignments)
    for xf in wb._cell_styles:
        if xf.alignmentId >= max_align:
            xf.alignmentId = 0
    wb.save(output_buffer)
