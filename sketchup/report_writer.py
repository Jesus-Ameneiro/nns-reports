"""
NNS Report Writer - fills MCC and CS templates with processed data.
"""

import base64
import io
import re
import zipfile
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date


# ---------------------------------------------------------------------------
# MCC FOOTER CONSTANTS
# Code owns the MCC Data sheet footer entirely.
# The template only needs to supply rows 1–31 (header + 18 data rows).
# Update these values when Trimble provides new contact information.
# ---------------------------------------------------------------------------

MCC_NOTE_GAP    = 5   # rows between last data row and the Nota row
MCC_CONTACT_GAP = 8   # rows between last data row and first contact line

MCC_FOOTER_NOTE = (
    "Nota: El presente documento contiene información confidencial "
    "y se proporciona exclusivamente dentro del marco de License Compliance."
)

# (text, bold) — one tuple per contact block line
MCC_CONTACT_LINES = [
    ("XXXXXXXXXX",                                    True),   # specialist name
    ("Especialista en Resolución",                    False),
    ("(xx) xxxx - xxxx",                              False),  # phone
    ("xxxxx@ruvixx.com",                              False),  # email
    ("",                                              False),
    ("425 Page Mill Rd, Suite 200, Palo Alto, 94306", False),  # address
]


# ---------------------------------------------------------------------------
# MCC CELL-EMBEDDED IMAGE (image1.png — 102×34 px compliance logo)
# Extracted from the MCC template's xl/media/image1.png.
# The template stores this image in cells A27 (LC Summary) and A43 (Data)
# using Excel's "Insert image in cell" rich-value feature (vm attribute).
# openpyxl strips that mechanism on read/write; we re-insert it as a
# standard floating OneCellAnchor drawing from this embedded PNG constant.
# If Trimble updates the logo, replace this value with the new base64.
# ---------------------------------------------------------------------------

_MCC_CELL_IMAGE_B64 = (
    'iVBORw0KGgoAAAANSUhEUgAAAGYAAAAiCAYAAACtFqwYAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAAEnQAABJ0Ad5mH3gAAA21SURBVGhD7Vp7sFxFmb9BeQkBU+DqioqsWGDKJSFzzpmZ/h595iYhV4OhxFwfqIuP4nrPOTM3El9F1mVYkUdk3d1yt1aQ8vGHDwziq1AMC0kgKCpayrolUi6lYoEG0WiABHLndG99fc6ZOTP3zs01BivGfFVdM9Pn6/769K+/7/t194yMHJb5yIKRkZEj8s/DcpDJwQ+OjStoYv+rJgreblpjRw8+PwRFwDiIAblQH2Oi4L0m8XfYZmBN5O1Jo+CTdqryokHdQ0zEUw5Ob3kqVktN5N1iY99KMbFv0sgzNglsJw7um078V9t2WwZ/yIodGVngynp98pPrqi8zbztj4aDO0yGzuqvdNP6Mp2L/ounEf9C2AmsT35rEN/LpvgtQrao1cbA7jf1/M29/+XMH+ziEZIFt1s/sJP6WNPEfNol3o4395w0qHSgpAJkBipn0zpiOgk+nkTctoJjYTwWUMjDdElddSSP/eybxVw32dahImvgb7bqqlVBu1wW2E3lvGtQ5UFLEzgXtkV4o6iT+BSbyf2qbMogsdBWgyPcyKCbxbSfxrI0Ca+OafN+VRpUr7Lolz+439ZcvadO/zk7JIgxSu64u77puUOdASZcGSo4wcf3UNPauT2PvSdusOS8oAZI6gGYA49lO07Mm9qwRcJLAOkCTYJudqqIdH39GZsoO9U6RWq12rNb65CAIThimU4i13b72R+bVLrfRJ51mMN6Jgz12qm5NUt2x92KoDeoMF+mvKP3SzvJzX33BOJx04uANJvaecK6ah6eyp7ikPwBMBo7kmaoDxsQ1RwrE06bjZR+1bf3MzGhvMkdHRxcT0WWI2EbEi4ioiYhfBoDvIOJWIvrI6OjoKYj4eiLaiIhXA0BzvAty1g8zv4KI5NkHEfG9SqnnI+J66VcpJeUtpTZdUUo1iFybS4non5n5pUR0ITNfhYhS3kpEL2TmNhFdUdhvjy8+ykTV800UXGrj+vLTg7ETEPEfEflKRL5GKV6jlKL89+WIdDVRmIf2HijMy88m0h9EDC8HoCu01ktnA8ZV2Cl4kY318Xsnl1Eaed+WpG4TyRuBMXGW+POSOjBirweMA1Lqqta2ajZNvB2d2I92xIuPt+uqL956oT6m7C3M/NowDK3W2hJRB5EMM8t3y6wtIj7BzD4i31DoIeJPJiYmjpSVXKxmRPxQ8RwAdtRqtdMB4OasX7ZKwW9kgssvnLf7Sm5b2u1CxL9DxNsbjYaVQkR3aa2fjYjfCENtmUVPTTPDq8r9MKr1xGxZNywhWwB+hVLqDAD6rdYyrtAi0v+VxzA+Pn4UIn2VWZ43pM39AFBsObrgdL+Y2H+/Sbwvmfjs082bVh6XRv6H0kl/t+QOE1WNES+QXJJUjGl6Nk0CASAHqOI8xJGExN9soupZtq2PMZP+lWkUfOXBtbVjCzsizHy+DIpIJqe/yIAReSdio4JIn5Hf2TP+YbvdFu8r93OVPJcCwL/UWj+PuTFR9IXIKaJ+fbmNTAKifjjTIauU+qIsUAFBwMoB2yZzE4YhENHjzGSYyCLgd8ayUDtSb9RPJcCfk4wvs//pkXzLAKAvyd8jH5v+sLXZXBPR64hwWhYOIk8ThW8sj2+GmGawwbZ8m05WHugkyy6QEGTiyivNZHCveI4UEwkD82zaXGY7ScWaqJKBMuVyzW9N4l0iXifAmNi/VRJlOlnd8r/txUeVbQ0CA8ApAD+ISP8DwA8Q8f2IeNYfAwwi/7JWGz1Fa30mAD9SgAxAHysvQET9BkTuyMTkHpNk9bhZvDYHZmulUjlSYj8iXu/qEC0DGkRsiT7UaSMX40d+tM58dmHD95efBEA/Li2QnfI+55577rMkXItdsQ/A26SuaDerdAQYCU1N36ZxZdok3rVPvGPJKbveuvQ5aVT5Txv7v3fhTcJay7OmtcTa1tnWxBXJO/9t4mVVB2YSJCbxH3G0UvY4UfV2Owcw+Yq6Yvny5SfJRKxcufK4lSvpNK318QB8wx8DDMAKCQkLEHlzb1Lwx5WKPln0pX8i/kSxmhHxUQk98kyAKXnMVgk5uY2XIeLDjGRdUXA/1uvjhLTDhV0OrUK+RnQlxOaJXNpNyAJAZCP26nX1KWZeL6AgYoqIuyVHlt9nVjHN2gaZdMeuJDw5muzdK14jBk1rWc1GlcvSSW9zmnjfS6Old5ho6bXTceU8866zjnMhMPFvMonfceRBQBZgJv0t+wImDMNq+XkhiHrTfgAzQqTfmdcZIt5bq6GWemF9iPzTzK6WMHareEZmqx+YxYt7Y2bmD2jJJUgdAkyB4BEENIBs68i/IKLTRC8DJct/a9asWYjI23OPMQD4BCLuBICOeCYifiGzPZytOSkDk8aekeJCW1LZ04mWvbvQE/pr2+1nyqlAty6phqYZPGCnMjbWJQoOGG+fwABodP30U9QF+w/M6DJE3pW1cwn4X7J6Wg1A073YD919CAD0hbIyMC53Id2nJfwBpkxkyBEWCcPh+0QnY39SepNMROch8lMCjBAcRBcKpexUSnlZ7/sERkJZ99glL+I5zoOeNJOBi8WDYhLfM83gIQfKAI2eDzB5uIH8UR8wRHqTMJucsc0bGK31MYj4zVKMv3vt2rXHAsCHZTKzxIs7G43GkqKv2XJM/ihjkoixAyab2O6Y6vX638jzzFvaR5RZ4/j44qOUopuy8WWg5O3+q5/GzwmMt6EbgspFzsPkNDnxf2cmA7eh6h7ovW/FiaZVu9O+U8jBQLt5ApMNOiyAKUvuMbK63UTOGxgRRN5QAANAjxHROYj4rYKWI+L20uTPBUyWM4BfpZHFW9x4cgp/c+ZZ5VXfA0YEgD+SLQ5nsyAc3QiUyf4CI2XKgXOTHP9320zVJs1UtZdTBssBAKbwGKXUDzdt2tS3WZwbGKwQ6d1ZkhcGRJ9DxD8Uqx0RN+SqxZ5oGDAL3KkE4uYw8xjLugvuHiJ1XqbWD07e51nCEDPannlaxsjofiL627z/PIQPBaa2IaPFQ4BJXP7ZbVqZ15j3qIUm9u7JNqKzgPKnAzMCgJ8sgEHEn4+OjvadXjPzlcyNWYERVofILpzlMV4mx63anBH5ueq+gBkJw/CNjJRqWfUZZZ4myjbFAHDXihUrTsw0e8BIe9nb5PbFdkcSfxHO5DSh6H9O2TcwWUiziX+Z0295ysTenhn6BxAYpfCqPB/IZDyFiBPFM0RcRKRvnd1jihXLl+fPJJR0YzwAfFdoed7VnMDkJwDf16xtnmP+QET/SkS78v2IAaB39OxmtgEgVIoez5ihC3vfAoDP5h4jv38tVDwfw3CZE5juRGc7e6cfB+8ZGsIOEDAAfD4AThcrHQB/D4A3KEX/jkj3ZBtFnQ4DhogUIufhrA+YjbkJ0SsB06PLWrszPqHe75Z6RrZh9nmt7HEY+cuu3jEz/Imc0xW2x8bGjoY6fU2OacQ2kd4LwK/OTh3wIZL9kDt60h8XwpCPZXYZmmP6JtoB8yPzrpXHmab3UdlEzno/s//A9MXZ1atXL0LkH/TCUZbMpU1eV7y45JAZwOTh7J5ee6ff0Tqj5/35gDdLWMzHtUXqzjnnnBcS6V84rxQiAfyrhmq4DWmIoSbkxxjFk0JLoP+pGDcpeg0BP0kgex9tGfgWYYrOjuK2q5N3Ad41bFF2Zf7AeD8zF9dOMYl/w/4Cg4hrG43lbhLCcNQqxZQ/mpEAZS8AwL9zkzsARnasot0hIgDtGARGBJGvEjsyudmhIf0ov1oYAIa2hOFyK+NC1HeKxyDyf8j4XFvJZYovLfp15AT4egHFAaMkyY8uHpMTZ6AfhNSwGkOLih6TsFY0Gq2OPheB7gu5YUWHFH1dyEWp337pNKvvz27nZpngPmD8B+XexiTe5/cJjJyVxf624t80QrHlU1YbU7idiLcy6ju01i/Ph1EGpnt5RxSuBoXfQKBHJTwQ8m4A/jYRXUJAXyLUWxD151etWuWYjpXJ7p5AhwBAtyDwZiZ9GyJeXOgUxdlAfbXmxm3MjdsI9MZGvXGqJv1FIt5GqO9E5C/IVUS3fznIrOszxRsIeTsj383Q+AdEvYIUb2PQdzDpu1Dx5YNXDwB8gcbwTlK0HYFvlw1x/mjGwhwxTTrNJP6NRo718/v9GSUHRq4I5vSY/KQ5jf17pyfrq7r7nryIW/u+f1KlUjlxRaVyoqzOgvvLSwsg+ct3wVmj1izUdX2mBl1rQGOJHK9IvYQrYUVjY2MnyD6nPOFSZFJER4pSaqHE/0EdKeV+hBiMnT52tJzf5cl/kVJr3J8vBgGVoxcJuVKkregV7VYjLiqzu560jyj6lSL2Su88U+xE5UiT+C3TDH7lrlAHJ3w+wDgducPxrzOt4AXdvt3LFODMHEBpU1aA0ec9pe9DxdmQnXcxebPcQooMgjLbeIbJ/rSbj+6+njsxzUog7KtY/d1LsSHAOHCcl7h/zPzMtII35zeX+yvDBum8aLDyIJSnb4zCvDpxsMHE3qPdvDMEmPzIRrzkc2aq+tLBvg7L0yB7J3xOo8rdLu9MOSD6gHF1cfCwnfImzMVzsIvDcmAkZ1DOJW3094vSpn+NSaqPm6j6UMbK/Bvl8NLE/tflxnKw/WF5+iQHppeUTRycZ+LqzeYi9RIzGVxnIu8DplXsCQ7Ln0kyujqYcE1r7AS517frK46uHpbD8lcj/w8FUOciCgTP7wAAAABJRU5ErkJggg=='
)
_MCC_CELL_IMAGE_BYTES = base64.b64decode(_MCC_CELL_IMAGE_B64)

# ---------------------------------------------------------------------------
# CS FOOTER CONSTANTS
# Code owns the CS Data sheet footer for all three templates (ARG, Q1, PAR).
# The template only needs to supply rows 1–23 (header + 12 data rows).
# Update these values when Trimble provides new contact information.
# ---------------------------------------------------------------------------

CS_NOTE_GAP    = 9   # rows from last data row to the Note row   (23 + 9  = 32)
CS_IMAGE_GAP   = 7   # rows from last data row to footer image   (0-based row 30 = Excel 31)
CS_MERGE_D_GAP = 5   # rows from last data row to the D:M merge  (23 + 5  = 28)

CS_FOOTER_NOTE = (
    "Note: This document contains confidential information and is "
    "provided exclusively within the framework of License Compliance."
)


# ---------------------------------------------------------------------------
# UTILITIES
# ---------------------------------------------------------------------------

def detect_template_type(wb):
    return 'MCC' if 'LC Summary' in wb.sheetnames else 'CS'

def detect_summary_sheet(wb):
    for name in ['LC Summary', 'Summary', 'New Template']:
        if name in wb.sheetnames:
            return name
    return wb.sheetnames[0]

def find_col_by_header(ws, header_row, header_name):
    for cell in ws[header_row]:
        if cell.value and str(cell.value).strip().lower() == header_name.strip().lower():
            return cell.column
    return None

def safe_set(ws, row, col, value):
    """Set cell value, skipping merged-cell slaves silently."""
    try:
        ws.cell(row=row, column=col).value = value
    except AttributeError:
        pass

def col_all_dash(ws, col_idx, data_start_row, data_end_row):
    for r in range(data_start_row, data_end_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v is not None and str(v).strip() not in ('', '-'):
            return False
    return True

def format_date(d):
    if d is None:
        return '-'
    if hasattr(d, 'strftime'):
        return d.strftime('%Y-%m-%d')
    return str(d)


# ---------------------------------------------------------------------------
# MCC TEMPLATE FILLER
# ---------------------------------------------------------------------------

def _write_mcc_footer(ws_data, last_data_row, img_bytes=None, img_width=102, img_height=34):
    """
    Write the MCC Data sheet footer at a position computed from last_data_row.
    Called for every run — regardless of machine count — so the footer is
    always correctly placed and never overwritten by machine data.

    Writes:
      • Nota row (merged A:H) at last_data_row + MCC_NOTE_GAP
      • Contact block lines starting at last_data_row + MCC_CONTACT_GAP
      • Cell-embedded image (MCC_CONTACT_LINES index 4 = gap row) restored
        as a floating OneCellAnchor image if img_bytes is supplied.

    img_bytes: raw PNG bytes of the rich-value cell image extracted from the
               template (xl/media/image1.png via wb._archive).  The template
               has this image in cells A27 (LC Summary) and A43 (Data) using
               Excel's "Insert image in cell" (rich-value) feature which
               openpyxl cannot preserve natively.  We re-insert it as a
               standard floating drawing so it appears in the same position.
    """
    from openpyxl.styles import Font, Alignment

    note_row    = last_data_row + MCC_NOTE_GAP
    contact_row = last_data_row + MCC_CONTACT_GAP

    # ── Note cell ─────────────────────────────────────────────────────────
    cell            = ws_data.cell(note_row, 1)
    cell.value      = MCC_FOOTER_NOTE
    cell.font       = Font(name='Calibri', size=9, italic=True, color='404040')
    cell.alignment  = Alignment(wrap_text=True, vertical='top')

    ws_data.merge_cells(
        start_row=note_row, start_column=1,
        end_row=note_row,   end_column=8,
    )

    # ── Contact block ──────────────────────────────────────────────────────
    for offset, (text, bold) in enumerate(MCC_CONTACT_LINES):
        cell           = ws_data.cell(contact_row + offset, 1)
        cell.value     = text if text else None
        cell.font      = Font(name='Calibri', size=9, bold=bold)
        cell.alignment = Alignment(wrap_text=False)

    # ── Restore cell-embedded image at the gap row (MCC_CONTACT_LINES[4]) ─
    # The template stores a logo in A43 (Data) using Excel's rich-value
    # cell-image feature (vm attribute).  openpyxl strips this on read.
    # We re-insert the extracted image bytes as a floating drawing anchored
    # at the same cell so the logo appears in the correct position.
    # Row height is set to 22.2pt to match LC Summary A27 (the same image
    # there has ht=22.2pt in the template — ensures the logo is fully visible).
    if img_bytes:
        from openpyxl.drawing.image import Image as _XLImg
        _img            = _XLImg(io.BytesIO(img_bytes))
        _img.width      = img_width
        _img.height     = img_height
        _img_row        = contact_row + 4
        ws_data.row_dimensions[_img_row].height = max(22.2, img_height * 0.75)
        ws_data.add_image(_img, f'A{_img_row}')


# ---------------------------------------------------------------------------
# TEMPLATE IMAGE EXTRACTION
# openpyxl strips Excel's "Insert image in cell" (rich-value / vm attribute)
# on read/write.  These helpers extract every vm= image from the raw template
# ZIP and re-insert them as standard floating OneCellAnchor drawings.
# ---------------------------------------------------------------------------

def _extract_template_images(raw_bytes):
    """
    Extract all rich-value (vm=) cell images from the raw template bytes.

    Returns
    -------
    vm_cell_images : dict  {(sheet_index_1based, cell_coord): (img_bytes, w_px, h_px)}
        One entry per vm= cell found in any sheet.
    """
    import struct as _struct

    def _png_dims(data):
        if data[:4] == b'\x89PNG':
            return (_struct.unpack('>I', data[16:20])[0],
                    _struct.unpack('>I', data[20:24])[0])
        return None, None

    media   = {}            # {filename: (bytes, w, h)}
    vm_map  = {}            # {vm_index_1based: filename}
    result  = {}            # {(sheet_idx, coord): (bytes, w, h)}

    try:
        with zipfile.ZipFile(io.BytesIO(raw_bytes)) as z:
            names = z.namelist()

            # ── Extract all media images ─────────────────────────────────
            for name in names:
                if name.startswith('xl/media/'):
                    fname = name.split('/')[-1]
                    data  = z.read(name)
                    w, h  = _png_dims(data)
                    media[fname] = (data, w, h)

            # ── Parse richValue rels: rId → filename ─────────────────────
            rv_rels = {}
            rels_path = 'xl/richData/_rels/richValueRel.xml.rels'
            if rels_path in names:
                rels_xml = z.read(rels_path).decode()
                for m in re.finditer(
                        r'Id="(rId\d+)"[^>]*Target="[^"]*?([^/"]+)"', rels_xml):
                    rv_rels[m.group(1)] = m.group(2)

            # ── Parse rdrichvalue.xml: vm index → rId (via LocalImageIdentifier) ─
            rv_xml_path = 'xl/richData/rdrichvalue.xml'
            if rv_xml_path in names:
                rv_xml = z.read(rv_xml_path).decode()
                # Each <rv> first <v> = LocalImageIdentifier (0-based index into sorted rIds)
                entries = re.findall(r'<rv[^>]*>.*?<v>(\d+)</v>', rv_xml, re.DOTALL)
                sorted_rids = sorted(rv_rels.keys(), key=lambda x: int(x[3:]))
                for i, img_id_str in enumerate(entries):
                    img_id = int(img_id_str)
                    if img_id < len(sorted_rids):
                        fname = rv_rels.get(sorted_rids[img_id])
                        if fname:
                            vm_map[i + 1] = fname   # vm is 1-based

            # ── Scan each sheet for cells with vm= attribute ─────────────
            for si in range(1, 10):
                sname = f'xl/worksheets/sheet{si}.xml'
                if sname not in names:
                    break
                content = z.read(sname).decode('utf-8')
                for row_m in re.finditer(
                        r'<row r="(\d+)"[^>]*>(.*?)</row>', content, re.DOTALL):
                    rb = row_m.group(2)
                    for cm in re.finditer(r'<c r="([^"]+)"[^>]*vm="(\d+)"', rb):
                        coord   = cm.group(1)
                        vm_idx  = int(cm.group(2))
                        fname   = vm_map.get(vm_idx)
                        if fname and fname in media:
                            result[(si, coord)] = media[fname]

    except Exception:
        pass   # template may not have rich-value images — skip silently

    return result


def _restore_header_images(ws, vm_cell_images, sheet_idx, max_row=5):
    """
    Re-insert header-area vm= images (rows 1–max_row) as floating drawings.
    Called after writing all data so we don't conflict with row operations.

    The original templates anchor these images at row 1 with a very small
    (1pt) row height; the images extend visually into the rows below.
    The underlying cells held '#VALUE!' as cached errors — cleared here
    so the worksheet doesn't show error text behind the floating images.
    """
    from openpyxl.drawing.image import Image as _XLImg
    for (si, coord), (img_bytes, w, h) in vm_cell_images.items():
        if si != sheet_idx:
            continue
        m = re.match(r'[A-Z]+(\d+)', coord)
        if not m or int(m.group(1)) > max_row:
            continue
        if img_bytes and w and h:
            # Clear the cached #VALUE! error from the original vm= cell.
            # Must reset data_type explicitly: openpyxl reads vm= error cells
            # as data_type='e'; setting .value=None must also clear the type
            # or the cell still serialises as <c t="e"><v>#VALUE!</v></c>.
            try:
                _ecell = ws.cell(
                    row=int(re.match(r'[A-Z]+(\d+)', coord).group(1)),
                    column=sum((ord(ch)-64)*26**i
                               for i,ch in enumerate(
                                   reversed(re.match(r'([A-Z]+)', coord).group(1))))
                )
                _ecell.value     = None
                _ecell.data_type = 'n'
            except Exception:
                pass
            _img        = _XLImg(io.BytesIO(img_bytes))
            _img.width  = w
            _img.height = h
            ws.add_image(_img, coord)


def fill_mcc(wb, rows, globals_data, case_ids, entity_name, country,
             vm_cell_images=None):
    ws_summary = wb['LC Summary']
    ws_data    = wb['Data']

    # ── Cell-embedded image: use the pre-embedded PNG constant ─────────────
    # The MCC template stores a logo in A27 (LC Summary) and A43 (Data)
    # using Excel's rich-value cell-image feature (vm attribute) which
    # openpyxl strips on read/write.  Use the pre-embedded constant bytes.
    _img_bytes = _MCC_CELL_IMAGE_BYTES

    # ---- LC Summary (footer stays at fixed template positions: 20, 23-26, 28) ----
    ws_summary['B8']  = ', '.join(case_ids)
    ws_summary['B9']  = entity_name
    ws_summary['A14'] = country
    ws_summary['B14'] = globals_data['total_machines']
    ws_summary['C14'] = globals_data['total_users']
    ws_summary['D14'] = globals_data['versions_str']
    ws_summary['E14'] = globals_data['total_events']
    ws_summary['F14'] = globals_data['total_licenses']
    ws_summary['G14'] = globals_data['years_of_use']
    ws_summary['H14'] = globals_data['period']
    ws_summary['B16'] = globals_data['total_machines']
    ws_summary['C16'] = globals_data['total_users']
    ws_summary['D16'] = globals_data['versions_str']
    ws_summary['E16'] = globals_data['total_events']
    ws_summary['F16'] = globals_data['total_licenses']
    ws_summary['G16'] = globals_data['years_of_use']

    # ── LC Summary: overwrite contact block with clean plain-text values ──
    # The MCC template has formula-like placeholders in rows 23–28 that
    # start with '=' (e.g. '=B2452(xx) xxxx - xxxx'). Excel evaluates them
    # as formulas and displays #VALUE!. Write clean strings from
    # MCC_CONTACT_LINES so Excel never sees a formula it can't resolve.
    from openpyxl.styles import Font as _Font
    _LC_CONTACT_START = 23
    for _offset, (_text, _bold) in enumerate(MCC_CONTACT_LINES):
        _cell       = ws_summary.cell(_LC_CONTACT_START + _offset, 1)
        _cell.value = _text if _text else None
        _cell.font  = _Font(name='Calibri', size=9, bold=_bold)

    # ── Restore cell-embedded image at A27 (LC Summary gap row) ───────────
    # Same image as in the Data sheet — always anchored at A27 in LC Summary
    # since the contact block here is at fixed rows 23–28.
    # Row height 22.2pt matches the original template (row 27 has ht=22.2pt,
    # the only row in LC Summary with a non-standard height).
    if _img_bytes:
        from openpyxl.drawing.image import Image as _XLImg2
        _img2        = _XLImg2(io.BytesIO(_img_bytes))
        _img2.width  = 102
        _img2.height = 34
        _lc_img_row  = _LC_CONTACT_START + 4   # = 27
        ws_summary.row_dimensions[_lc_img_row].height = 22.2
        ws_summary.add_image(_img2, f'A{_lc_img_row}')

    # Check for COMPUTER DOMAIN column in summary header row
    SUMMARY_HEADER_ROW = 13
    comp_domain_col_summary = find_col_by_header(ws_summary, SUMMARY_HEADER_ROW, 'COMPUTER DOMAIN')

    # ---- Data sheet ----
    DATA_HEADER_ROW  = 13
    DATA_START_ROW   = 14
    TEMPLATE_DATA_ROWS = 18   # template has 18 pre-bordered rows (14-31)

    col_map = {}
    for cell in ws_data[DATA_HEADER_ROW]:
        if cell.value:
            col_map[str(cell.value).strip()] = cell.column

    mcc_col_order = [
        ('Active MAC',             'active_mac'),
        ('# Licenses',             'license_count'),
        ('Products',               'product'),
        ('First Event',            'first_event'),
        ('Last Event',             'last_event'),
        ('Event Types',            'event_type'),
        ('Computer Domains',       'computer_domain'),
        ('Version',                'version'),
        ('IP Country',             'ip_country'),
        ('Hostname',               'hostname'),
        ('Username',               'username'),
        ('Client Email Addresses', 'client_email'),
    ]

    n_rows = len(rows)
    n_template_cols = max(col_map.values()) if col_map else 8

    # Template layout constants
    TEMPLATE_LAST_DATA_ROW = DATA_START_ROW + TEMPLATE_DATA_ROWS - 1  # row 31

    last_data_row = DATA_START_ROW + n_rows - 1
    excess        = TEMPLATE_DATA_ROWS - n_rows

    # ── Clear old template footer zone ────────────────────────────────────
    # Must happen BEFORE writing data rows so that data written into rows
    # 32+ is not subsequently wiped. Merges cleared first to avoid the
    # AttributeError raised when assigning to a merged-cell slave.
    # Only remove footer-zone merges (rows > TEMPLATE_LAST_DATA_ROW).
    # Preserves header merges (A1:B4, H1:I4) that must not be destroyed.
    _footer_start = TEMPLATE_LAST_DATA_ROW + 1
    for _m in [str(m) for m in ws_data.merged_cells.ranges
               if m.min_row >= _footer_start]:
        ws_data.merged_cells.remove(_m)

    for r in range(TEMPLATE_LAST_DATA_ROW + 1, TEMPLATE_LAST_DATA_ROW + 25):
        for c in range(1, n_template_cols + 1):
            try:
                ws_data.cell(r, c).value = None
            except AttributeError:
                pass  # residual merged slave — safe to skip after range.clear()

    # ── Write data rows ────────────────────────────────────────────────────
    # Copy template styles to any row beyond the 18 pre-bordered template rows.
    for idx, row in enumerate(rows):
        r = DATA_START_ROW + idx
        if idx >= TEMPLATE_DATA_ROWS:
            _copy_row_style(ws_data, DATA_START_ROW, r, n_template_cols)
        for header, field in mcc_col_order:
            col_idx = col_map.get(header)
            if col_idx is None:
                continue
            val = row.get(field, '-')
            if val is None:
                val = '-'
            # Write dates as datetime objects to preserve Excel date formatting
            if field in ('first_event', 'last_event') and val != '-':
                try:
                    from datetime import datetime, date as date_type
                    if isinstance(val, date_type) and not isinstance(val, datetime):
                        val = datetime(val.year, val.month, val.day)
                except Exception:
                    val = format_date(val)
            safe_set(ws_data, r, col_idx, val)

    # ── Row count adjustment ───────────────────────────────────────────────
    if excess > 0:
        # Fewer machines than the 18 template rows: delete the empty surplus.
        # After deletion the footer is written at the correct position below.
        ws_data.delete_rows(DATA_START_ROW + n_rows, excess)

    # ── Write footer at the computed position ─────────────────────────────
    # Always done in code — works for 1 machine or 1 000+ machines.
    _write_mcc_footer(ws_data, last_data_row, img_bytes=_img_bytes)

    data_end_row = last_data_row

    # Fix stale cell-level hyperlink refs after delete_rows
    for row in ws_data.iter_rows():
        for cell in row:
            if cell.hyperlink and cell.hyperlink.ref:
                correct_ref = f'{get_column_letter(cell.column)}{cell.row}'
                if cell.hyperlink.ref != correct_ref:
                    cell.hyperlink.ref = correct_ref

    # ---- Column deletion ----
    # ── Restore Data header vm= images BEFORE column deletion ─────────────
    # Same ordering requirement as fill_cs: header images must be added to
    # ws_data BEFORE delete_cols so _fix_image_anchors_after_col_deletion
    # correctly shifts them (e.g. H1 → G1 when Computer Domains col G deleted).
    if vm_cell_images:
        _restore_header_images(ws_data, vm_cell_images, sheet_idx=2)

    # Column deletion — also shifts the header images we just added above
    cols_to_delete = []
    for header in ['Computer Domains', 'Client Email Addresses']:
        col_idx = col_map.get(header)
        if col_idx and col_all_dash(ws_data, col_idx, DATA_START_ROW, data_end_row):
            cols_to_delete.append(col_idx)

    if col_map.get('Computer Domains') in cols_to_delete and comp_domain_col_summary:
        ws_summary.delete_cols(comp_domain_col_summary)

    for col_idx in sorted(cols_to_delete, reverse=True):
        ws_data.delete_cols(col_idx)
        _fix_image_anchors_after_col_deletion(ws_data, col_idx - 1)

    # ── Restore Summary header vm= images ─────────────────────────────────
    # Summary is unaffected by Data column deletion — order safe here.
    if vm_cell_images:
        _restore_header_images(ws_summary, vm_cell_images, sheet_idx=1)

    return wb


# ---------------------------------------------------------------------------
# CS TEMPLATE FILLER
# ---------------------------------------------------------------------------

def _fix_merged_cells_after_row_deletion(ws, first_deleted_row, count=1):
    """
    openpyxl's delete_rows() shifts cell contents but does NOT update merged
    cell range coordinates. Rebuild the registry, decrementing any row number
    >= first_deleted_row by count (number of rows deleted).
    """
    old_ranges = [
        (mc.min_row, mc.min_col, mc.max_row, mc.max_col)
        for mc in ws.merged_cells.ranges
    ]
    ws.merged_cells.ranges.clear()
    for (min_row, min_col, max_row, max_col) in old_ranges:
        # Drop merges entirely within the deleted range
        if min_row >= first_deleted_row and max_row < first_deleted_row + count:
            continue
        # Keep merges entirely above the deleted range unchanged
        if max_row < first_deleted_row:
            ws.merge_cells(start_row=min_row, start_column=min_col,
                           end_row=max_row, end_column=max_col)
            continue
        # Shift rows at or below the first deleted row
        if min_row >= first_deleted_row:
            min_row -= count
        if max_row >= first_deleted_row:
            max_row -= count
        if min_row < 1 or max_row < 1 or min_row > max_row:
            continue
        ws.merge_cells(start_row=min_row, start_column=min_col,
                       end_row=max_row, end_column=max_col)


def _fix_image_anchors_after_row_deletion(ws, deleted_row_0based):
    """Shift image anchors >= deleted_row_0based down by 1 (0-based rows)."""
    for img in ws._images:
        try:
            anchor = img.anchor
            if hasattr(anchor, '_from'):
                if anchor._from.row >= deleted_row_0based:
                    anchor._from.row -= 1
            if hasattr(anchor, 'to') and anchor.to is not None:
                if anchor.to.row >= deleted_row_0based:
                    anchor.to.row -= 1
        except Exception:
            pass


def _fix_image_anchors_after_rows_deletion(ws, first_deleted_0based, count):
    """Shift image anchors >= first_deleted_0based down by count (0-based rows)."""
    for img in ws._images:
        try:
            anchor = img.anchor
            if hasattr(anchor, '_from'):
                if anchor._from.row >= first_deleted_0based:
                    anchor._from.row -= count
            if hasattr(anchor, 'to') and anchor.to is not None:
                if anchor.to.row >= first_deleted_0based:
                    anchor.to.row -= count
        except Exception:
            pass


def _fix_merged_cells_after_col_deletion(ws, deleted_col_1based):
    """
    openpyxl's delete_cols() does not update merged cell col coordinates.
    Decrement any col index >= deleted_col_1based by 1.
    Merges entirely within one cell or that span the deleted col are dropped.
    """
    old_ranges = [
        (mc.min_row, mc.min_col, mc.max_row, mc.max_col)
        for mc in ws.merged_cells.ranges
    ]
    ws.merged_cells.ranges.clear()
    for (min_row, min_col, max_row, max_col) in old_ranges:
        # Drop merges entirely on the deleted column
        if min_col == deleted_col_1based and max_col == deleted_col_1based:
            continue
        # Shift cols >= deleted col left by 1
        if min_col >= deleted_col_1based:
            min_col -= 1
        if max_col >= deleted_col_1based:
            max_col -= 1
        if min_col > max_col:
            continue
        ws.merge_cells(start_row=min_row, start_column=min_col,
                       end_row=max_row, end_column=max_col)


def _fix_image_anchors_after_col_deletion(ws, deleted_col_0based):
    """
    Shift image anchor cols >= deleted_col_0based left by 1 (0-based cols).

    Handles two anchor types that openpyxl uses:
      • String anchor (e.g. 'K1') — created by ws.add_image(img, 'K1')
      • AnchorMarker object      — created when loading an existing workbook
    String anchors do NOT have a _from attribute so the old isinstance check
    silently skipped them, leaving images at the wrong column after deletion.
    """
    from openpyxl.utils import get_column_letter, column_index_from_string
    import re as _re_col

    for img in ws._images:
        try:
            anchor = img.anchor
            if isinstance(anchor, str):
                # Simple string anchor like 'K1' — parse col letter, shift if needed
                m = _re_col.match(r'([A-Z]+)(\d+)', anchor)
                if m:
                    col_1based = column_index_from_string(m.group(1))
                    col_0based  = col_1based - 1
                    if col_0based >= deleted_col_0based:
                        new_col_letter = get_column_letter(col_1based - 1)
                        img.anchor = f'{new_col_letter}{m.group(2)}'
            elif hasattr(anchor, '_from'):
                if anchor._from.col >= deleted_col_0based:
                    anchor._from.col -= 1
                if hasattr(anchor, 'to') and anchor.to is not None:
                    if anchor.to.col >= deleted_col_0based:
                        anchor.to.col -= 1
        except Exception:
            pass


def _copy_row_style(ws, src_row, dst_row, max_col):
    """Copy cell styles (font, border, fill, alignment, number_format) from src_row to dst_row."""
    from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, numbers
    import copy
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        try:
            dst.font      = copy.copy(src.font)
            dst.border    = copy.copy(src.border)
            dst.fill      = copy.copy(src.fill)
            dst.alignment = copy.copy(src.alignment)
            if src.number_format and src.number_format != 'General':
                dst.number_format = src.number_format
        except Exception:
            pass


def _write_cs_footer(ws_data, last_data_row, img_bytes=None,
                     img_width=None, img_height=None):
    """
    Write the CS Data sheet footer at a position computed from last_data_row.
    Called for every run so the footer is always correctly placed.

    Writes:
      • D:M merge at last_data_row + CS_MERGE_D_GAP  (structural, no text)
      • Footer image at last_data_row + CS_IMAGE_GAP if img_bytes provided.
        The CS template stores this as a rich-value cell image (vm attribute)
        which openpyxl strips; we re-insert it as a standard floating drawing.
      • Note text + A:J merge at last_data_row + CS_NOTE_GAP
    """
    from openpyxl.styles import Font, Alignment

    note_row    = last_data_row + CS_NOTE_GAP
    image_row   = last_data_row + CS_IMAGE_GAP   # 0-based anchor row
    merge_d_row = last_data_row + CS_MERGE_D_GAP

    # ── D:M structural merge ──────────────────────────────────────────────
    ws_data.merge_cells(
        start_row=merge_d_row, start_column=4,
        end_row=merge_d_row,   end_column=13,
    )

    # ── Footer image ──────────────────────────────────────────────────────
    # Add the dynamically-extracted vm= image at the computed position.
    # (Old: reposition existing floating images — this no longer applies since
    #  all CS Data sheet images are rich-value cells, not floating drawings.)
    if img_bytes and img_width and img_height:
        from openpyxl.drawing.image import Image as _XLImg
        _img        = _XLImg(io.BytesIO(img_bytes))
        _img.width  = img_width
        _img.height = img_height
        _excel_img_row = image_row + 1  # convert 0-based to 1-based for add_image
        ws_data.add_image(_img, f'A{_excel_img_row}')

    # ── Note text + merge A:J ─────────────────────────────────────────────
    cell           = ws_data.cell(note_row, 1)
    cell.value     = CS_FOOTER_NOTE
    cell.font      = Font(name='Calibri', size=9, italic=True, color='404040')
    cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws_data.merge_cells(
        start_row=note_row, start_column=1,
        end_row=note_row,   end_column=10,
    )


def fill_cs(wb, rows, globals_data, case_ids, entity_name, country, vm_cell_images=None):
    summary_name   = detect_summary_sheet(wb)
    ws_summary     = wb[summary_name]
    ws_data        = wb['Data']
    vm_cell_images = vm_cell_images or {}

    # ── Pre-extract footer and header images from vm_cell_images ──────────
    # CS Summary A37: small footer logo (always at original row 37 in template)
    _cs_sum_footer_img = vm_cell_images.get((1, 'A37'))   # (bytes, w, h) or None
    # CS Data A31: footer logo (at row 31 = last_data_row+7+1 for 12 machines)
    _cs_dat_footer_img = (vm_cell_images.get((2, 'A31'))
                          or vm_cell_images.get((2, 'A30'))
                          or None)

    # ---- Determine if Computer Domain row should be deleted ----
    # Template row 12 = 'Computer Domain'.
    # When no domain data is present, delete that row entirely so all
    # subsequent rows shift up by 1, matching the expected output exactly.
    all_comp_domains = set()
    for row in rows:
        cd = row.get('computer_domain', '-')
        if cd and cd != '-':
            for d in cd.split(','):
                d = d.strip()
                if d:
                    all_comp_domains.add(d)

    delete_domain_row = len(all_comp_domains) == 0
    COMP_DOMAIN_ROW   = 12

    if delete_domain_row:
        # 1. Delete the row
        ws_summary.delete_rows(COMP_DOMAIN_ROW)
        # 2. Fix merged cell coordinates (openpyxl bug: they don't shift automatically)
        _fix_merged_cells_after_row_deletion(ws_summary, COMP_DOMAIN_ROW)
        # 3. Fix stale cell-level hyperlink refs in summary sheet after row deletion
        for summ_row in ws_summary.iter_rows():
            for cell in summ_row:
                if cell.hyperlink and cell.hyperlink.ref:
                    correct_ref = f'{get_column_letter(cell.column)}{cell.row}'
                    if cell.hyperlink.ref != correct_ref:
                        cell.hyperlink.ref = correct_ref
        # 4. Fix image anchor rows in summary sheet (deleted row = 12, 0-based = 11)
        _fix_image_anchors_after_row_deletion(ws_summary, COMP_DOMAIN_ROW - 1)
        # Row mappings after deletion (each was original_row - 1)
        ver_row  = 12   # Version row  (was 13)
        yofu_row = 13   # Years of Use (was 14)
        vals_row = 18   # Machines/Users/Events values (was 19)
        lic_row  = 30   # Licensed copies value row (was 31)
        # 3. Update price formula: it referenced C31 which is now C30
        for r in range(25, 35):
            cell = ws_summary.cell(r, 1)
            if cell.value and 'C31' in str(cell.value):
                cell.value = str(cell.value).replace('C31', 'C30')
                break
    else:
        ws_summary['B12'] = ', '.join(sorted(all_comp_domains))
        ver_row  = 13
        yofu_row = 14
        vals_row = 19
        lic_row  = 31

    # ---- Summary sheet: fill all data fields ----
    ws_summary['B9']  = ', '.join(case_ids)
    ws_summary['B10'] = country
    ws_summary['B11'] = entity_name

    # Version string → B{ver_row} (master of B:E merge)
    # Total Versions  → G{ver_row} (free cell, not merged)
    # Single-year versions are written as integers to match Excel native format
    versions_val = globals_data['versions_str']
    try:
        if ',' not in str(versions_val):
            versions_val = int(versions_val)
    except (ValueError, TypeError):
        pass
    ws_summary.cell(ver_row, 2).value = versions_val
    ws_summary.cell(ver_row, 7).value = globals_data['total_versions']

    # Years of Use → B{yofu_row}
    # Period       → D{yofu_row} (master of D:G merge)
    ws_summary.cell(yofu_row, 2).value = globals_data['years_of_use']
    ws_summary.cell(yofu_row, 4).value = globals_data['period']

    # Machines / Users / Events numeric values
    # B, D, G are masters of their merged spans in the values row
    safe_set(ws_summary, vals_row, 2, globals_data['total_machines'])
    safe_set(ws_summary, vals_row, 4, globals_data['total_users'])
    safe_set(ws_summary, vals_row, 7, globals_data['total_events'])

    # Licensed copies (feeds the price formula)
    safe_set(ws_summary, lic_row, 3, globals_data['total_licenses'])

    # ---- Data sheet: header rows ----
    safe_set(ws_data, 6, 2,  ', '.join(case_ids))
    safe_set(ws_data, 7, 2,  entity_name)
    safe_set(ws_data, 8, 2,  country)
    # DATE: label at col 10, current date at col 11 with mm-dd-yy format
    from datetime import datetime as _dt
    safe_set(ws_data, 7, 10, 'DATE:')
    today_cell = ws_data.cell(7, 11)
    today_cell.value = _dt.today().replace(hour=0, minute=0, second=0, microsecond=0)
    today_cell.number_format = 'mm-dd-yy'

    DATA_HEADER_ROW    = 11
    DATA_START_ROW     = 12
    TEMPLATE_DATA_ROWS = 12   # 12 pre-bordered rows (12-23) in the blank template

    col_map = {}
    for cell in ws_data[DATA_HEADER_ROW]:
        if cell.value:
            col_map[str(cell.value).strip()] = cell.column

    cs_col_order = [
        ('Products',               'product'),
        ('Version',                'version'),
        ('Event Types',            'event_type'),
        ('Active MAC',             'active_mac'),
        ('# Licenses',             'license_count'),
        ('First Event',            'first_event'),
        ('Last Event',             'last_event'),
        ('Computer Domains',       'computer_domain'),
        ('IP Country',             'ip_country'),
        ('Hostname',               'hostname'),
        ('Username',               'username'),
        ('Client Email Addresses', 'client_email'),
    ]

    n_rows = len(rows)

    TEMPLATE_LAST_DATA_ROW = DATA_START_ROW + TEMPLATE_DATA_ROWS - 1  # row 23
    excess        = TEMPLATE_DATA_ROWS - n_rows
    last_data_row = DATA_START_ROW + n_rows - 1
    n_template_cols = max(col_map.values()) if col_map else 12

    # ── Clear old template footer zone ────────────────────────────────────
    # Must happen BEFORE writing data rows so data written to rows 24+
    # is not subsequently wiped.
    # IMPORTANT: only remove footer-zone merges (rows > TEMPLATE_LAST_DATA_ROW).
    # Header merges (e.g. A1:B5, K1:L5) must be preserved.
    _footer_start = TEMPLATE_LAST_DATA_ROW + 1
    for _m in [str(m) for m in ws_data.merged_cells.ranges
               if m.min_row >= _footer_start]:
        ws_data.merged_cells.remove(_m)

    for r in range(TEMPLATE_LAST_DATA_ROW + 1, TEMPLATE_LAST_DATA_ROW + 30):
        for c in range(1, n_template_cols + 1):
            try:
                ws_data.cell(r, c).value = None
            except AttributeError:
                pass  # residual merged slave

    # ── Write machine data rows ────────────────────────────────────────────
    # Dates as datetime objects (preserves template formatting).
    # Copy styles from template row 12 for any row beyond the 12-row zone.
    STYLE_SRC_ROW = DATA_START_ROW
    for idx, row in enumerate(rows):
        r = DATA_START_ROW + idx
        if idx >= TEMPLATE_DATA_ROWS:
            _copy_row_style(ws_data, STYLE_SRC_ROW, r, n_template_cols)
        for header, field in cs_col_order:
            col_idx = col_map.get(header)
            if col_idx is None:
                continue
            val = row.get(field, '-')
            if val is None:
                val = '-'
            if field in ('first_event', 'last_event') and val != '-':
                try:
                    from datetime import datetime, date as date_type
                    if isinstance(val, date_type) and not isinstance(val, datetime):
                        val = datetime(val.year, val.month, val.day)
                except Exception:
                    val = format_date(val)
            safe_set(ws_data, r, col_idx, val)

    # Apply correct date format to date columns
    date_fmt = 'yyyy\\-mm\\-dd'
    date_fields = ['First Event', 'Last Event']
    for df in date_fields:
        col_idx = col_map.get(df)
        if col_idx:
            for r in range(DATA_START_ROW, DATA_START_ROW + n_rows):
                ws_data.cell(r, col_idx).number_format = date_fmt

    # ── Row count adjustment ───────────────────────────────────────────────
    if excess > 0:
        # Fewer machines than the 12 template rows: delete empty surplus.
        # No merge/anchor fix needed — _write_cs_footer repositions everything.
        ws_data.delete_rows(DATA_START_ROW + n_rows, excess)

    # ── Write footer at the computed position ─────────────────────────────
    # Always done in code — works for 1 machine or 1 000+ machines.
    if _cs_dat_footer_img:
        _dat_img_bytes, _dat_img_w, _dat_img_h = _cs_dat_footer_img
    else:
        _dat_img_bytes, _dat_img_w, _dat_img_h = None, None, None
    _write_cs_footer(ws_data, last_data_row,
                     img_bytes=_dat_img_bytes,
                     img_width=_dat_img_w,
                     img_height=_dat_img_h)

    data_end_row = last_data_row

    # Fix stale cell-level hyperlink refs after delete_rows
    for data_row in ws_data.iter_rows():
        for cell in data_row:
            if cell.hyperlink and cell.hyperlink.ref:
                correct_ref = f'{get_column_letter(cell.column)}{cell.row}'
                if cell.hyperlink.ref != correct_ref:
                    cell.hyperlink.ref = correct_ref

    # ── Restore Data header vm= images BEFORE column deletion ─────────────
    # CRITICAL ORDER: header images must be inserted into ws_data BEFORE any
    # column deletion.  openpyxl's _fix_image_anchors_after_col_deletion will
    # then shift the image anchors correctly (e.g. K1→J1 after deleting col H).
    # If we added images AFTER deletion, we'd place them at the wrong column.
    if vm_cell_images:
        _restore_header_images(ws_data, vm_cell_images, sheet_idx=2)

    # ---- Column deletion ----
    cols_to_delete = []
    for header in ['Computer Domains', 'Client Email Addresses']:
        col_idx = col_map.get(header)
        if col_idx and col_all_dash(ws_data, col_idx, DATA_START_ROW, data_end_row):
            cols_to_delete.append(col_idx)

    for col_idx in sorted(cols_to_delete, reverse=True):
        ws_data.delete_cols(col_idx)
        # Fix image anchor columns (openpyxl doesn't update them after delete_cols)
        # This also shifts any header images we just added above.
        _fix_image_anchors_after_col_deletion(ws_data, col_idx - 1)  # convert to 0-based
        # Fix merged cell col ranges (openpyxl doesn't update them after delete_cols)
        _fix_merged_cells_after_col_deletion(ws_data, col_idx)

    # ── Restore CS Summary A37 footer logo ────────────────────────────────
    if _cs_sum_footer_img:
        from openpyxl.drawing.image import Image as _XLImgCS
        _sf_bytes, _sf_w, _sf_h = _cs_sum_footer_img
        _sf_row = 36 if delete_domain_row else 37
        try:
            _sf_cell = ws_summary.cell(_sf_row, 1)
            _sf_cell.value     = None
            _sf_cell.data_type = 'n'
        except Exception:
            pass
        _sf_img         = _XLImgCS(io.BytesIO(_sf_bytes))
        _sf_img.width   = _sf_w
        _sf_img.height  = _sf_h
        ws_summary.add_image(_sf_img, f'A{_sf_row}')

    # ── Restore Summary header vm= images (Summary not affected by Data col deletion) ─
    if vm_cell_images:
        _restore_header_images(ws_summary, vm_cell_images, sheet_idx=1)

    return wb


# ---------------------------------------------------------------------------
# MAIN ENTRY POINT
# ---------------------------------------------------------------------------

def fill_template(template_wb, rows, globals_data, case_ids, entity_name, country,
                  raw_bytes=None):
    """
    raw_bytes: the raw .xlsx file bytes, used to extract rich-value (vm=)
               cell images before openpyxl strips them.  Pass the template
               file bytes (read before loading with openpyxl) for full
               image fidelity.  If None, only the pre-embedded MCC footer
               constant is available as a fallback.
    """
    wb            = template_wb
    template_type = detect_template_type(wb)

    # Extract all vm= cell images from the raw template bytes
    vm_cell_images = {}
    if raw_bytes:
        vm_cell_images = _extract_template_images(raw_bytes)

    if template_type == 'MCC':
        wb = fill_mcc(wb, rows, globals_data, case_ids, entity_name, country,
                      vm_cell_images=vm_cell_images)
    else:
        wb = fill_cs(wb, rows, globals_data, case_ids, entity_name, country,
                     vm_cell_images=vm_cell_images)
    return wb, template_type


def patch_and_save(wb, output_buffer):
    """Save workbook, patching style alignment corruption from delete_cols."""
    max_align = len(wb._alignments)
    for xf in wb._cell_styles:
        if xf.alignmentId >= max_align:
            xf.alignmentId = 0
    wb.save(output_buffer)
